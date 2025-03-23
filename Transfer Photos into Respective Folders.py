# Assumes photo number on the last row #
########################################
import openpyxl
import os
import os.path
import pandas as pd
import shutil
import re

########################################################
# Defining Variables, make sure everything is correct! #
########################################################
photoformatjpg = ".jpg" #code assumes only jpg, heic and png photos will be used
photoformatheic = ".heic"
photoformatpng = ".png"
Photofolder = "Photos" #name of folder containing all photos
excelformat = ".xlsx" # code assumes excel workbook format is .xlsx

cwd = os.getcwd()
listdir = os.listdir(cwd)
filtered_list = [x for x in listdir if re.search(excelformat,x)]
cwdphotofolder = cwd + "/" + Photofolder + "/" # Main folder where all photos are located
phot_dir = os.listdir(cwdphotofolder) #Lisitng and sieving out subfolders located in main folder

##################################################################
# Everything should be contained within this singular "for" loop #
##################################################################
for lanka in range(len(phot_dir)):
	workbook = filtered_list[lanka] #name of workbook with species name and photo number
	#####################################
	# Folder will be TXXX_Genus species #
	#####################################
	tag = "Tag" #Column name containing Tree Tag
	Sci = "Scientific.Name" #Column name containing Scientific name

	# cont. Defining Variables #
	cwd = os.getcwd()
	wb = openpyxl.load_workbook(workbook) 
	sheet = wb.active
	lastcol = sheet.max_column
	alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", 
					"R", "S", "T","U", "V", "W", "X", "Y", "Z"]
	column_with_photo_number = alphabet[lastcol-1] # Code assumes photo range located in last column


	print("It is better to do one’s own dharma, even though imperfectly")
	print("than to do another’s dharma, even though perfectly.") 
	print("By doing one’s innate duties, a person does not incur sin.\nBhagavad Gita 18.47")

	for lakshmana in range(sheet.max_row-1):
		print("\nRow " + str(lakshmana+1))
		table = pd.read_excel(workbook)

		######################################
		# Extacting Tag value of tree/animal #
		######################################

		tag_index = table.columns.get_loc(tag)
		cell_object_tag = sheet[str(alphabet[tag_index]) + str(lakshmana +2)].value

		##################################################
		# Extacting Scientific name value of tree/animal #
		##################################################

		Sci_index = table.columns.get_loc(Sci)
		cell_obj_sci = sheet[str(alphabet[Sci_index]) + str(lakshmana +2)].value


		#####################################################
		# Creating folder using template TXXX_Genus species #
		#####################################################
		filename = str(cell_object_tag) + "_" + (cell_obj_sci)
		try:
			os.makedirs(filename)
			print(filename + " created")
		except FileExistsError:
			pass

		#################################
		# Extacting image number values #
		#################################
		lansium = sheet[str(column_with_photo_number) + str(lakshmana+2)].value # If photo is in column B, change it to sheet["B3"]

		####################################
		# Adding photo numbers into a list #
		####################################

		########################################################
		# Substitute all spaces for nothing e.g " 40" ==> "40" #
		########################################################
		x = []
		try:
			split = lansium.split(',') #If there are no photos, move on to next row
		except AttributeError:
			print("Row " + str(lakshmana+1) + " has no Photos")
			continue

		for i in range(len(split)):
			x.append(split[i].replace(" ", ""))

		p = []
		###################################
		# How many digits are there in x? #
		###################################
		for i in range(len(x)): 
			z = x[i].split("-")
			digits = len(list(z[0])) # how many digits in photo number?
			z0 = int(z[0]) # this value repeats itself alot, best to define it
			

			######################################
			# If there is only one value e.g 998 #
			######################################
			if len(z) == 1:
				matsya = str(z0).zfill(digits)
				p.append(matsya)
				continue

			######################################################
			# If second value more than first value e.g. 996-998 #
			######################################################
			elif int(z[1]) -z0 > 0:
				for i in range(int(z[1]) -z0-1):
					if int(z[i]) - int(z[i+1]) != -1:
						kurma = str(int(z[i]) + i+1).zfill(digits)
						z.insert(0, str(kurma))

			##########################################################
			# If photo number goes beyond the threshold e.g. 998-004 #
			##########################################################
			elif int(z[1]) - z0 < 0:
				rnd = int(z[1])
				vamana = []
				for i in range(digits):
					vamana.append("9")
				max_number = int(''.join(vamana))
				for i in range(max_number - z0):
					if max_number - int(z[0]) > 0:
						z.insert(0, str(int(z[0]) +1).zfill(digits)) # DO NOT SUBSTITUTE int(z[0])!!!
				for i in range(rnd):
					a = int(z[-1]) - (i* 1)
					b = str(a).zfill(digits)
					z.insert(0, b) 
			for q in range(len(z)):
				p.append(z[q])

		#####################
		# remove duplicates #
		#####################
		res = []
		for value in p:
			if value not in res:
				res.append(value)
		sorted_res = sorted(res)
		
		print("Photo Numbers: " + lansium)

		#######################################
		# Moving photos to respective folders #
		#######################################
		for canarium in range(len(sorted_res)):
			photographer = cwdphotofolder + phot_dir[lanka]+ "/"

			##############################################
			# Defining copy, paste and renaming function #
			##############################################
			def onerepulic(photoformat):
				imagenumber = "IMG_" + sorted_res[canarium] + photoformat
				source = photographer + imagenumber
				destination = cwd + "/" + filename + "/"
				rename_format = phot_dir[lanka] +"("+imagenumber+")"+photoformat
				try:
					if os.path.exists(destination+"/"+imagenumber):
						pass
					else:
						shutil.copy(source, destination)
						print(imagenumber + " has been moved to " + filename)
						os.rename(destination+"/"+imagenumber, destination + rename_format)
						print(imagenumber + " has been renamed to " + rename_format)
				except FileNotFoundError:
					print("It appears " + imagenumber +" dosen't exist")
					pass
				except shutil.Error:
					print(imagenumber + " already exists in the destination folder!")
					exit()
				except FileExistsError:
					print("\n"+imagenumber + " cannot be renamed since")
					print(rename_format + " already exists in you directory")
					print("Python cannot rename a file to a name that already exists, please delete the file and try again.")
					exit()

			###################################
			# Transfer Photos with jpg format #
			###################################					
			if os.path.isfile(photographer +"IMG_" + sorted_res[canarium] + ".jpg") == True:
				onerepulic(".jpg")
				
			###################################
			# Transfer Photos with heic format #
			###################################
			elif os.path.isfile(photographer + "IMG_" + sorted_res[canarium] + ".heic") == True:
				onerepulic(".heic")
				
			###################################
			# Transfer Photos with png format #
			###################################
			elif os.path.isfile(photographer + "IMG_" + sorted_res[canarium] + ".png") == True:
				onerepulic(".png")